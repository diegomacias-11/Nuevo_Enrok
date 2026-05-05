import copy
from datetime import datetime
from io import BytesIO
from decimal import Decimal
from pathlib import Path
from xml.sax.saxutils import escape

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PyPDF2 import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.enums import TA_JUSTIFY
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


_POPPINS_REGISTERED = False


def _register_poppins_fonts():
    global _POPPINS_REGISTERED
    if _POPPINS_REGISTERED:
        return
    fonts_dir = settings.BASE_DIR / "static" / "fonts"
    font_files = {
        "Poppins": "Poppins-Regular.ttf",
        "Poppins-SemiBold": "Poppins-SemiBold.ttf",
        "Poppins-Bold": "Poppins-Bold.ttf",
    }
    for font_name, file_name in font_files.items():
        font_path = fonts_dir / file_name
        if font_path.exists():
            pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
    _POPPINS_REGISTERED = True


def _active(post_data, name):
    return post_data.get(name) == "on"


def _format_date(value):
    if not value:
        return timezone.localdate().strftime("%d-%m-%Y")
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d-%m-%Y")
    except Exception:
        return value


def _format_value(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    return str(value)


def _paragraph_text(value):
    return escape(value or "").replace("\n", "<br/>")


def _ordered_columns(post_data):
    ordered = [value for value in (post_data.get("columnas_orden") or "").split(",") if value]
    selected = post_data.getlist("columnas")
    if ordered:
        return [value for value in ordered if value in selected]
    return selected


def _safe_filename(value, default):
    name = (value or "").strip() or default
    invalid = '<>:"/\\|?*'
    for char in invalid:
        name = name.replace(char, "-")
    return name[:120].strip() or default


def _excel_value(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return value
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    return str(value)


def _numeric_value(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    return None


def build_pdf_report(rows, fields, post_data, filename="reporte.pdf"):
    _register_poppins_fonts()
    field_map = {field["name"]: field["label"] for field in fields}
    numeric_fields = {field["name"] for field in fields if field.get("is_numeric")}
    selected_columns = _ordered_columns(post_data)
    total_columns = [
        column
        for column in post_data.getlist("columnas_totales")
        if column in selected_columns and column in numeric_fields
    ]
    title_filename = _safe_filename(post_data.get("pdf_titulo"), "reporte")
    if _active(post_data, "pdf_bloque_fecha"):
        date_suffix = _format_date(post_data.get("pdf_fecha"))
        filename = f"{title_filename}_{date_suffix}.pdf" if date_suffix else f"{title_filename}.pdf"
    else:
        filename = f"{title_filename}.pdf"

    template_path = settings.BASE_DIR / "static" / "img" / "MEMBRETE.pdf"
    pagesize = letter
    template_reader = None
    template_page = None
    if template_path.exists():
        try:
            template_reader = PdfReader(str(template_path))
            template_page = template_reader.pages[0]
            pagesize = (float(template_page.mediabox.width), float(template_page.mediabox.height))
        except Exception:
            template_reader = None
            template_page = None

    buffer = BytesIO()
    page_width, page_height = pagesize
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        leftMargin=page_width * 0.03,
        rightMargin=page_width * 0.03,
        topMargin=page_height * 0.11,
        bottomMargin=page_height * 0.18,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Poppins-Bold",
        fontSize=16,
        leading=17,
        alignment=1,
        textColor=colors.HexColor("#1f2a3d"),
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Heading2"],
        fontName="Poppins-Bold",
        fontSize=14,
        leading=16,
        alignment=1,
        textColor=colors.HexColor("#334155"),
    )
    body_style = ParagraphStyle(
        "ReportBody",
        parent=styles["BodyText"],
        fontName="Poppins",
        fontSize=12,
        leading=14,
        textColor=colors.HexColor("#2b313f"),
    )
    paragraph_style = ParagraphStyle(
        "ReportParagraph",
        parent=body_style,
        alignment=TA_JUSTIFY,
    )
    small_style = ParagraphStyle(
        "ReportSmall",
        parent=styles["BodyText"],
        fontName="Poppins",
        fontSize=12,
        leading=14,
        textColor=colors.HexColor("#5f6f82"),
    )
    footer_style = ParagraphStyle(
        "ReportFooter",
        parent=body_style,
        alignment=TA_JUSTIFY,
    )
    table_header_style = ParagraphStyle(
        "ReportTableHeader",
        parent=body_style,
        fontName="Poppins-Bold",
        fontSize=7,
        leading=8,
        textColor=colors.HexColor("#1f2a3d"),
        alignment=1,
    )

    elements = []
    title_text = post_data.get("pdf_titulo") or ""
    subtitle_text = post_data.get("pdf_subtitulo") or ""
    date_text = _format_date(post_data.get("pdf_fecha") or "")

    if _active(post_data, "pdf_bloque_titulo") or _active(post_data, "pdf_bloque_fecha"):
        page_width = pagesize[0] - doc.leftMargin - doc.rightMargin
        side_width = 92
        center_width = max(120, page_width - (side_width * 2))
        title_block = Paragraph(title_text, title_style) if _active(post_data, "pdf_bloque_titulo") else ""
        date_block = Paragraph(date_text, small_style) if _active(post_data, "pdf_bloque_fecha") else ""
        heading = Table([["", title_block, date_block]], colWidths=[side_width, center_width, side_width])
        heading.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (1, 0), (1, 0), "CENTER"),
                    ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        elements.append(heading)

    if _active(post_data, "pdf_bloque_subtitulo") and subtitle_text:
        elements.append(Paragraph(subtitle_text, subtitle_style))

    if _active(post_data, "pdf_bloque_parrafo") and post_data.get("pdf_parrafo"):
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(_paragraph_text(post_data.get("pdf_parrafo")), paragraph_style))

    if _active(post_data, "pdf_bloque_tabla") and selected_columns:
        elements.append(Spacer(1, 12))
        table_data = [
            [Paragraph(field_map.get(column, column), table_header_style) for column in selected_columns]
        ]
        totals = {column: 0 for column in total_columns}
        has_total_values = {column: False for column in total_columns}
        for row in rows:
            row_data = []
            for column in selected_columns:
                raw_value = getattr(row, column, "")
                numeric_value = _numeric_value(raw_value)
                if column in totals and numeric_value is not None:
                    totals[column] += numeric_value
                    has_total_values[column] = True
                row_data.append(Paragraph(_format_value(raw_value), body_style))
            table_data.append(
                row_data
            )
        if total_columns:
            table_data.append(
                [
                    totals[column] if column in totals and has_total_values[column] else ""
                    for column in selected_columns
                ]
            )
        available_width = pagesize[0] - doc.leftMargin - doc.rightMargin
        col_widths = [available_width / len(selected_columns)] * len(selected_columns)
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef6fa")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1f2a3d")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#aebed2")),
            ("FONTNAME", (0, 0), (-1, 0), "Poppins-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Poppins"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        if total_columns:
            table_styles.extend(
                [
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#ddeaf4")),
                    ("FONTNAME", (0, -1), (-1, -1), "Poppins-Bold"),
                    ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#1f2a3d")),
                ]
            )
        table.setStyle(
            TableStyle(
                table_styles
            )
        )
        elements.append(table)

    if _active(post_data, "pdf_bloque_pie") and post_data.get("pdf_pie"):
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(_paragraph_text(post_data.get("pdf_pie")), footer_style))

    doc.build(elements)
    content_pdf = buffer.getvalue()
    buffer.close()

    if template_reader and template_page:
        content_reader = PdfReader(BytesIO(content_pdf))
        writer = PdfWriter()
        for page in content_reader.pages:
            base = copy.copy(template_page)
            if base.mediabox != page.mediabox:
                base.mediabox = page.mediabox
            base.merge_page(page)
            writer.add_page(base)
        output = BytesIO()
        writer.write(output)
        pdf = output.getvalue()
        output.close()
        return pdf, filename

    return content_pdf, filename


def build_excel_report(rows, fields, post_data):
    field_map = {field["name"]: field["label"] for field in fields}
    numeric_fields = {field["name"] for field in fields if field.get("is_numeric")}
    selected_columns = _ordered_columns(post_data)
    total_columns = [
        column
        for column in post_data.getlist("columnas_totales")
        if column in selected_columns and column in numeric_fields
    ]
    title = _safe_filename(post_data.get("excel_titulo_archivo"), "reporte")
    date_suffix = _format_date(post_data.get("excel_fecha_archivo")) if _active(post_data, "excel_incluir_fecha_archivo") else ""
    filename = f"{title}_{date_suffix}.xlsx" if date_suffix else f"{title}.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Hoja 1"

    include_headers = _active(post_data, "excel_incluir_encabezados")
    include_count = _active(post_data, "excel_incluir_conteo")

    header_fill = PatternFill("solid", fgColor="EEF6FA")
    total_fill = PatternFill("solid", fgColor="DDEAF4")
    border_color = "AEBED2"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    current_row = 1
    if include_headers:
        col_offset = 1
        if include_count:
            cell = worksheet.cell(row=current_row, column=1, value="Conteo")
            cell.font = Font(name="Poppins", bold=True, color="1F2A3D")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            col_offset = 2
        for col_index, column in enumerate(selected_columns, start=1):
            cell = worksheet.cell(row=current_row, column=col_index + col_offset - 1, value=field_map.get(column, column))
            cell.font = Font(name="Poppins", bold=True, color="1F2A3D")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        current_row += 1

    totals = {column: 0 for column in total_columns}
    has_total_values = {column: False for column in total_columns}
    for row_number, row in enumerate(rows, start=1):
        col_offset = 1
        if include_count:
            cell = worksheet.cell(row=current_row, column=1, value=row_number)
            cell.font = Font(name="Poppins", color="2B313F")
            cell.border = thin_border
            col_offset = 2
        for col_index, column in enumerate(selected_columns, start=1):
            raw_value = getattr(row, column, None)
            value = _excel_value(raw_value)
            cell = worksheet.cell(row=current_row, column=col_index + col_offset - 1, value=value)
            cell.font = Font(name="Poppins", color="2B313F")
            cell.border = thin_border
            numeric_value = _numeric_value(raw_value)
            if column in totals and numeric_value is not None:
                totals[column] += numeric_value
                has_total_values[column] = True
        current_row += 1

    if total_columns and selected_columns:
        col_offset = 1
        if include_count:
            cell = worksheet.cell(row=current_row, column=1, value="")
            cell.font = Font(name="Poppins", bold=True, color="1F2A3D")
            cell.fill = total_fill
            cell.border = thin_border
            col_offset = 2
        for col_index, column in enumerate(selected_columns, start=1):
            value = totals[column] if column in totals and has_total_values[column] else ""
            cell = worksheet.cell(row=current_row, column=col_index + col_offset - 1, value=value)
            cell.font = Font(name="Poppins", bold=True, color="1F2A3D")
            cell.fill = total_fill
            cell.border = thin_border

    if include_count:
        worksheet.column_dimensions["A"].width = 10

    col_offset = 2 if include_count else 1
    for col_index, column in enumerate(selected_columns, start=1):
        values = [str(field_map.get(column, column))]
        values.extend(str(_excel_value(getattr(row, column, ""))) for row in rows[:50])
        width = min(max(len(value) for value in values) + 2, 42)
        worksheet.column_dimensions[worksheet.cell(row=1, column=col_index + col_offset - 1).column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), filename
